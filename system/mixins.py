import collections
import datetime
import os

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.paginator import Paginator
from django.http import HttpRequest, QueryDict, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework.decorators import action
from rest_framework.response import Response

from constants import DOWNLOAD_JOB_LIMIT
from system.models import DownloadJob
import urllib.parse
import tempfile
import logging

from system.services import unique_id

logger = logging.getLogger(__name__)


class DownloadMixin:
    """
    Mixin for excel download of the ListViewSet.
    """
    model = None
    column = None
    file_name = "Download"

    def __init__(self, *args, **kwargs):
        super(DownloadMixin, self).__init__(*args, **kwargs)
        self.filter_dict = []
        self.export_excel = True
        self.request_data = None

    def get_queryset(self):
        return self.model.objects.all()

    def process_data(self, request_object, user=None):
        self.request_data = request_object
        request = HttpRequest()
        request.method = 'GET'
        params = urllib.parse.urlencode(request_object)
        request.query_params = request_object
        query_dict = QueryDict(params)
        request.GET = query_dict
        request.user = user
        self.request = request
        return self.get_excelsheet(self.filter_queryset(self.get_queryset()))

    def obj_to_list(self, qs):
        data = qs.values()
        return data

    def update_column(self, qs):
        if not self.column:
            data = self.obj_to_list(qs)
            keys = data[0].keys() if data else []
            self.column = [{"data": key, "title": key.replace('_', ' ').upper()}
                           for key in keys]
        return

    def get_excelsheet(self, qs):
        count = qs.count()
        self.update_column(qs[:1])
        paginate_by = DOWNLOAD_JOB_LIMIT
        paginator = Paginator(qs, paginate_by)
        num_page = paginator.num_pages
        u_id = str(unique_id())
        temp_path = "temp_file" + u_id + ".xlsx"
        temp_path = tempfile.gettempdir() + "/" + temp_path
        download_id = self.request_data.get('download_id', None)
        wb = self.create_file()
        wb.save(temp_path)
        row_number = 1
        try:
            download_job = DownloadJob.objects.get(id=int(download_id))
        except Exception as e:
            logger.info(f"{e}")
            return False

        download_job.data_count = count
        download_job.save()
        try:
            for i in range(1, num_page + 1):
                page = paginator.page(i)
                queryset = page.object_list
                data = self.obj_to_list(queryset)
                context = {
                    'items': data,
                    'count': count,
                }
                wb = self.write_file(context, row_number, temp_path)
                wb.save(temp_path)
                row_number = row_number + paginate_by
                complete_to = round((i / num_page) * 90, 2)
                download_job.complete_percentage = complete_to
                download_job.save()
        except Exception as e:
            logger.error(str(e))
            return False

        try:
            file_name = str(self.file_name)
        except:
            file_name = "Report"

        file_url = self.close(temp_path, file_name)
        data = {
            "file": file_url,
            "count": count
        }
        return data

    def create_file(self):
        try:
            excel_dict = []
            for col in self.column:
                excel_dict.append({'field': col['data'], 'column': col['title']})

            columns = [obj['column'] for obj in excel_dict]

            wb = Workbook()
            ws = wb.create_sheet("data sheet", 0)
            ws.append(columns)
            red_font = Font(bold=True)
            for cell in ws["1:1"]:
                cell.font = red_font
            return wb

        except Exception as e:
            logger.error(str(e))
            return False

    def write_file(self, data, row_number, temp_path):
        try:
            wb = openpyxl.load_workbook(temp_path)
            ws = wb.active
            excel_dict = []
            for col in self.column:
                excel_dict.append({'field': col['data'], 'column': col['title']})

            items = list(data['items'])
            rows = []
            for item in items:
                record = collections.OrderedDict()
                for obj in excel_dict:
                    column_name = obj.get('field', None)
                    record[column_name] = item[column_name]
                rows.append(record)

            replace_empty = True
            row_num = (row_number - 1)

            for key, row in enumerate(rows):
                row_num += 1
                col_num = 0

                for key, value in row.items():
                    if 'style_bold' in key:
                        col_num -= 1
                    elif not value and value != 0:
                        if replace_empty:
                            ws.cell(row=row_num + 1, column=col_num + 1, value='-----')
                        else:
                            ws.cell(row=row_num + 1, column=col_num + 1, value='')
                    elif isinstance(value, datetime.datetime):
                        naive = value.replace(tzinfo=None)
                        ws.cell(row=row_num + 1, column=col_num + 1, value=naive)
                    elif isinstance(value, datetime.date):
                        ws.cell(row=row_num + 1, column=col_num + 1, value=value)
                    else:
                        ws.cell(row=row_num + 1, column=col_num + 1, value=value)

                    col_num += 1
            return wb
        except Exception as e:
            logger.error(str(e))
            return False

    def close(self, temp_path, filename):
        try:
            try:
                filename = str(filename) + '.xlsx'
            except:
                filename = "Report.xlsx"
            wb = openpyxl.load_workbook(temp_path)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="' + filename + '"'
            wb.save(response)
            attachment = SimpleUploadedFile(filename, response.getvalue(), content_type='application/vnd.ms-excel')
            os.remove(temp_path)
            download_job = DownloadJob.objects.get(id=int(self.request_data.get('download_id', None)))
            download_job.complete_percentage = 99
            download_job.completed_at = timezone.now()
            download_job.save()
            return attachment
        except Exception as e:
            logger.error(str(e))
            return False

    @action(methods=['get', ], detail=False)
    def download(self, request):
        data = request.GET.dict()
        from system.download import get_download_type
        u = DownloadJob(
            request_data=data,
            download_type=get_download_type(self.__class__)
        )
        u.save()
        from system.tasks import execute_download
        task = execute_download.delay(u.id)
        u.task_id = task.id
        u.save()
        return Response()
